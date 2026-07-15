#!/usr/bin/env python3
"""Genera el portal HTML desde el Excel maestro de oportunidades.

Uso local:
    python scripts/generar_portal.py

El script conserva el diseño de index.html y reemplaza solamente:
- el arreglo JavaScript ``opportunities``;
- la fecha visible de actualización;
- la fecha de referencia usada para clasificar convocatorias.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DEFAULT_EXCEL = Path("datos/oportunidades_financiacion.xlsx")
DEFAULT_HTML = Path("index.html")
SHEET_NAME = "Oportunidades"

MONTHS_ES = (
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
)


def clean(value: Any) -> str:
    """Convierte una celda a texto limpio, preservando saltos como espacios."""
    if value is None:
        return ""
    return " ".join(str(value).replace("\r", " ").replace("\n", " ").split())


def normalized(value: Any) -> str:
    """Normaliza texto para comparaciones tolerantes a acentos y mayúsculas."""
    text = unicodedata.normalize("NFKD", clean(value).casefold())
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def split_multi(value: Any) -> list[str]:
    """Divide campos multivalor separados por punto y coma."""
    return [item.strip() for item in clean(value).split(";") if item.strip()]


def parse_date(value: Any, tipo_fecha: Any = None) -> date | None:
    """Convierte la fecha de Excel; las convocatorias continuas no tienen cierre."""
    if normalized(tipo_fecha) == "continua":
        return None
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Fecha no reconocida: {value!r}")


def format_date_es(value: date | None, tipo_fecha: Any = None) -> str:
    if value is None:
        if normalized(tipo_fecha) == "continua":
            return "Convocatoria continua"
        return "Sin fecha confirmada"
    return f"{value.day} de {MONTHS_ES[value.month - 1]} de {value.year}"


def format_number(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)):
        if float(value).is_integer():
            return f"{int(value):,}".replace(",", ".")
        return f"{float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return clean(value)


def format_amount(minimum: Any, maximum: Any, currency: Any) -> str:
    """Construye una etiqueta de monto sin asumir que ambos extremos son numéricos."""
    cur = clean(currency)
    min_text = format_number(minimum)
    max_text = format_number(maximum)
    min_norm = normalized(minimum)
    max_norm = normalized(maximum)

    if not min_text and not max_text:
        return "Monto no especificado"
    if min_norm in {"sin limite", "sin límite"} or max_norm in {"sin limite", "sin límite"}:
        base = min_text if min_text and "limite" not in min_norm else max_text
        return f"Desde {cur} {base}, sin límite".strip()
    if min_text and max_text:
        if min_text == max_text:
            return f"{cur} {max_text}".strip()
        if normalized(minimum) in {"0", "0.0"}:
            return f"Hasta {cur} {max_text}".strip()
        return f"{cur} {min_text} – {max_text}".strip()
    if max_text:
        return f"Hasta {cur} {max_text}".strip()
    return f"Desde {cur} {min_text}".strip()


def combine_eligibility(row: dict[str, Any]) -> str:
    parts = []
    principal = clean(row.get("Elegibilidad"))
    if principal:
        parts.append(principal)

    extras = (
        ("Solicitante", row.get("Tipo de solicitante")),
        ("Ubicación elegible", row.get("Ubicación elegible")),
        ("Ciudadanía", row.get("Ciudadanía elegible")),
    )
    for label, value in extras:
        text = clean(value)
        if text:
            parts.append(f"{label}: {text}.")
    return " ".join(parts) or "Consultar las bases oficiales."


def read_opportunities(excel_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise KeyError(f"No existe la hoja {SHEET_NAME!r}. Hojas disponibles: {wb.sheetnames}")

    ws = wb[SHEET_NAME]
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [clean(value) for value in next(rows)]
    except StopIteration as exc:
        raise ValueError("El Excel está vacío.") from exc

    required = {
        "ID", "Publicar", "Título de la oportunidad", "Financiador",
        "Fecha límite", "Tipo de fecha", "Tipo de financiamiento",
        "Resumen", "Enlace oficial",
    }
    missing = sorted(required.difference(headers))
    if missing:
        raise KeyError(f"Faltan columnas obligatorias: {', '.join(missing)}")

    opportunities: list[dict[str, Any]] = []
    for excel_row_number, values in enumerate(rows, start=2):
        row = dict(zip(headers, values))
        if normalized(row.get("Publicar")) != "si":
            continue
        if not clean(row.get("Título de la oportunidad")):
            continue

        deadline = parse_date(row.get("Fecha límite"), row.get("Tipo de fecha"))
        url = clean(row.get("Enlace oficial"))
        if not url:
            raise ValueError(f"Fila {excel_row_number}: falta el enlace oficial.")

        profiles = split_multi(row.get("Perfil profesional"))
        if not profiles:
            profiles = ["No especificado"]

        opportunity = {
            "id": len(opportunities) + 1,
            "code": clean(row.get("ID")),
            "title": clean(row.get("Título de la oportunidad")),
            "funder": clean(row.get("Financiador")),
            "funderType": clean(row.get("Tipo de financiador")),
            "country": clean(row.get("País del financiador")),
            "deadline": deadline.isoformat() if deadline else None,
            "deadlineLabel": format_date_es(deadline, row.get("Tipo de fecha")),
            "deliveryType": clean(row.get("Tipo de entrega")),
            "manualStatus": clean(row.get("Estado manual")),
            "calculatedStatus": clean(row.get("Estado calculado")),
            "amount": format_amount(
                row.get("Monto mínimo"), row.get("Monto máximo"), row.get("Moneda")
            ),
            "duration": clean(row.get("Duración")) or "No especificada",
            "type": clean(row.get("Tipo de financiamiento")) or "Otro",
            "career": profiles,
            "applicantType": clean(row.get("Tipo de solicitante")),
            "area": clean(row.get("Área temática principal")),
            "keywords": split_multi(row.get("Palabras clave")),
            "summary": clean(row.get("Resumen")),
            "eligibility": combine_eligibility(row),
            "notes": clean(row.get("Notas de la convocatoria")),
            "internalObservation": clean(row.get("Observaciones internas")),
            "url": url,
            "pivotUrl": clean(row.get("Enlace Pivot-RP")),
        }
        opportunities.append(opportunity)

    if not opportunities:
        raise ValueError("No se encontraron filas válidas con Publicar = Sí.")
    return opportunities


def update_html(html_path: Path, opportunities: list[dict[str, Any]]) -> None:
    html = html_path.read_text(encoding="utf-8")
    data = json.dumps(opportunities, ensure_ascii=False, indent=2)

    data_pattern = re.compile(
        r"const opportunities\s*=\s*\[.*?\];\s*\nconst today\s*=.*?;",
        flags=re.DOTALL,
    )
    replacement = (
        f"const opportunities={data};\n"
        "const today=new Date();today.setHours(0,0,0,0);"
    )
    html, count = data_pattern.subn(lambda _: replacement, html, count=1)
    if count != 1:
        raise RuntimeError(
            "No fue posible localizar el bloque 'const opportunities' en index.html. "
            "No cambies ese nombre en la plantilla."
        )

    generated = date.today()
    date_text = f"{generated.day} de {MONTHS_ES[generated.month - 1]} de {generated.year}"
    html = re.sub(
        r"Actualizado desde el Excel maestro:\s*[^<]+",
        f"Actualizado desde el Excel maestro: {date_text}",
        html,
        count=1,
    )

    html_path.write_text(html, encoding="utf-8", newline="\n")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--html", type=Path, default=DEFAULT_HTML)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        opportunities = read_opportunities(args.excel)
        update_html(args.html, opportunities)
    except Exception as exc:  # Mensaje compacto para GitHub Actions
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print(f"Portal actualizado: {args.html} ({len(opportunities)} oportunidades publicadas).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
